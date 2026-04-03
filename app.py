import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import io
import traceback
import os
from openai import OpenAI

st.set_page_config(page_title="Excel AI Uzmanı", layout="wide", page_icon="🟢")

st.markdown("""
<style>
    /* Modern Tasarım */
    .stApp {
        background-color: #fcefeƒ;
    }
    
    h1 {
        color: #047857; /* Koyu yeşil - Excel hissiyatı */
        font-weight: 800;
        margin-bottom: -10px;
    }
    
    .stButton button {
        background-color: #10B981 !important;
        color: white !important;
        border-radius: 8px;
        font-weight: bold;
    }
    .stButton button:hover {
        background-color: #059669 !important;
    }
    
    .stDownloadButton button {
        background-color: #3B82F6 !important;
        color: white !important;
        border-radius: 8px;
        font-weight: bold;
    }
    .stDownloadButton button:hover {
        background-color: #2563EB !important;
    }
    
    /* İndir Butonu Konteynırı */
    div[data-testid="stVerticalBlock"] div:has(div.stDownloadButton) {
        text-align: center;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("<h3 style='color: #047857; font-weight: bold;'>Not: Bu Uygulama İş Güvenliği Uzmanı Fatih AKDENİZ tarafından geliştirilmiştir.</h3>", unsafe_allow_html=True)

# --- API KEY ALANI ---
try:
    api_key = st.secrets["OPENAI_API_KEY"]
except:
    api_key = os.environ.get("OPENAI_API_KEY")



# --- APP STATE & FILES ---
WORKSPACE_FILE = "workspace_excel.xlsx"

if "messages" not in st.session_state:
    st.session_state["messages"] = []

if "file_ready" not in st.session_state:
    st.session_state["file_ready"] = False

# --- DOSYA YÜKLEME / OLUŞTURMA (DOĞRUDAN EKRANIN ORTASI) ---
if not st.session_state["file_ready"]:
    st.info("👋 Başlamak için aşağıdaki seçeneklerden birini kullanın:")
    col_up, col_new = st.columns(2)
    
    with col_up:
        st.subheader("📁 Mevcut Dosyanızı Yükleyin")
        uploaded_file = st.file_uploader("Sadece .xlsx dosyaları:", type=["xlsx"])
        if uploaded_file:
            with open(WORKSPACE_FILE, "wb") as f:
                f.write(uploaded_file.getbuffer())
            st.session_state["file_ready"] = True
            st.session_state["messages"] = [{"role": "assistant", "content": "Dosyanızı inceledim. Hangi hücrelere excel formülü yazalım veya nasıl düzenleyelim?"}]
            st.rerun()

    with col_new:
        st.subheader("📝 Sıfırdan Boş Excel Yarat")
        st.write("Kendiniz bir test dosyası üzerinden işlemi deneyimleyin.")
        if st.button("Boş Excel İle Başla (Örnek Tablo)", use_container_width=True):
            wb = openpyxl.Workbook()
            ws = wb.active
            # Başlıklar
            basliklar = ["Ürün Adi", "Aylik Satis", "Birim Fiyat", "KDV Orani", "Net Satis"]
            ws.append(basliklar)
            
            # Biraz örnek dummy veri
            ws.append(["Elma", 120, 15, 0.20, ""])
            ws.append(["Armut", 80, 20, 0.20, ""])
            ws.append(["Kavun", 45, 30, 0.10, ""])
            
            wb.save(WORKSPACE_FILE)
            st.session_state["file_ready"] = True
            st.session_state["messages"] = [{"role": "assistant", "content": "Size 3 satırlık meyve satışları içeren test tablosu hazırladım. Örneğin: **'Net satış sütununu satış ile birim fiyatı çarparak KDV dahil bir excel formülü yazdır.'** diyebilirsiniz!"}]
            st.rerun()
            
else:
    # --- ÇALIŞMA ALANI ---
    col_preview, col_actions = st.columns([5, 2])
    
    with col_preview:
        st.subheader("📊 Tablonun Son Hali")
        try:
            # Sadece okuma/basma amacıyla pandas:
            # Pandas ile formüller (örn: "=B2*C2") ekranda "NaN" veya "text" gözükebilir ama Excel'de düzgün açılacaktır.
            # Ekranda null görünebilecek formül hücreleri için endişelenmeye gerek yok, dilersek string olarak gösterebiliriz (pandas engine calisiyor)
            df_preview = pd.read_excel(WORKSPACE_FILE, engine="openpyxl")
            st.dataframe(df_preview, use_container_width=True, height=200)
        except Exception as e:
            st.error(f"Önizleme gösterilemedi: {e}")
            
    with col_actions:
        st.subheader("💾 İndirme & Ayarlar")
        
        with open(WORKSPACE_FILE, "rb") as f:
            st.download_button(
                label="📥 İndir (Gerçek Formüllerle)",
                data=f,
                file_name="Ai_Duzenlenmis_Umarim_Son.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        st.divider()
        if st.button("❌ Bu Dosyayı Kapat", use_container_width=True):
            st.session_state["file_ready"] = False
            st.session_state["messages"] = []
            st.rerun()
            
    st.divider()
    
    # Sohbet Alanı
    for msg in st.session_state["messages"]:
        if msg["role"] == "user":
            st.chat_message("user").write(msg["content"])
        else:
            st.chat_message("assistant", avatar="🟢").write(msg["content"])
            if "code" in msg:
                with st.expander("🛠 AI'ın Yazdığı OpenPyXL Kodunu İncele"):
                    st.code(msg["code"], language="python")

    user_query = st.chat_input("Hücre rengi değiştirme, Matematiksel Excel formülü hesaplamaları, Sütun silme vb. söyleyin.")
    
    if user_query:
        if not api_key:
            st.error("Sistem Yapılandırma Hatası: OpenAI API Anahtarı sunucuya eklenmemiş.")
            st.stop()
            
        st.session_state["messages"].append({"role": "user", "content": user_query})
        st.chat_message("user").write(user_query)
        
        with st.chat_message("assistant", avatar="🟢"):
            with st.spinner("AI Dosyayı Analiz Ediyor, Formülleri işliyor... 🚀"):
                # Excel bağlamını alma
                try:
                    wb_temp = openpyxl.load_workbook(WORKSPACE_FILE)
                    ws_temp = wb_temp.active
                    max_row = ws_temp.max_row
                    max_col = ws_temp.max_column
                    
                    sample_data = []
                    for row in ws_temp.iter_rows(min_row=1, max_row=min(max_row, 15), values_only=True):
                        sample_data.append(row)
                        
                    context = f"""
                    MEVCUT EXCEL DOSYASI (SEKME: {ws_temp.title}):
                    - Satır Sayısı: {max_row} | Sütun Sayısı: {max_col}
                    - İlk {min(max_row, 15)} Satır Verisi:
                    {sample_data}
                    
                    ÖNEMLİ: Tablonun gerçek başlıklarının (Header) hangi satırda olduğunu bu verilere bakarak analiz et. Başlık satırından sonraki satırlarda işlem yap!
                    """
                except Exception as e:
                    context = f"Veriler okunamadı. Hata: {e}"

                system_prompt = f"""
                Sen DÜNYA ÇAPINDA BAŞARILI, İLERİ DÜZEY BİR EXCEL ve PYTHON 'openpyxl' UZMANISIN! 
                Müşteriler, bir işlemi mükemmel yapan, karmaşık durumları dahi hatasız çözen aşırı profesyonel sistemlere ihtiyaç duyarlar. Sen tam olarak busun.
                Aksi belirtilmedikçe tüm işlemleri 'openpyxl' ile MEVCUT excel dosyası üzerinde gerçekleştireceksin. 
                
                GÖREVİN: Kullanıcının isteğini anlayıp, bunu excel katmanında GERÇEK FORMÜLLER (=EĞER, =VLOOKUP vb.), profesyonel renklendirmeler ve sütun organizasyonları ile native (kalıcı) şekilde uygulayan, HATASIZ çalışan bir Python kodu yazmak.
                
                ALTIN KURALLAR (HATA YAPARSAN SİSTEM ÇÖKER):
                1. Kodu asla markdown (```python ... ```) içine alma! YALNIZCA SAF PYTHON KODU DÖNDÜR.
                2. Çalışmaya `import openpyxl` ve `from openpyxl.styles import PatternFill, Font, Alignment, Border, Side` vb. ile başla.
                3. Dosyayı `wb = openpyxl.load_workbook('{WORKSPACE_FILE}')` ile aç ve `ws = wb.active` ile aktif sayfasını seç. 
                4. Kullanıcı Matematiksel, İstatistiksel veya Koşullu bir işlem isterse bunu Excel Formülü (`=SUM(A2:B2)`, `=IF(...)`) olarak hücreye yaz ki dosya indirildiğinde Excel'de değiştirilebilir olsun! 
                5. MergedCell (Birleşik Hücre) Koruması: Hücreleri döngüyle tararken: `if type(ws.cell(r, c)).__name__ == 'MergedCell': continue` KESİNLİKLE UYGULA! (İSG tabloları vb. için hayatidir).
                6. ZAMAN SINIRLAMASI: Binlerce satırlık tablolarda `max_row` bazen milyonlar gösterebilir. Satırların boş olup olmadığını test et. Atıyorum üst üste 10 satır tamamen None ise `break` ile döngüyü erkenden bitir (Optimizasyon dehası ol).
                7. EN SONA MUTLAKA `wb.save('{WORKSPACE_FILE}')` YAZ!
                8. HİÇBİR ZAMAN dosya indirme tetikleme vs. yapma, `print` dahi KULLANMA.
                9. Buton (Macro, Form Control) EKLENEMEZ! Bunları açıkça reddetme, hücreyi butonmuş gibi (örn boyayıp) geçiştirebilirsin ama sahte openpyxl widget modülleri icat etme.
                10. Sütun harflerini bulmak için `from openpyxl.utils import get_column_letter` kullanmayı unutma.
                
                EXCEL UZMANLIĞI KALİTE TESTİ:
                - Kullanıcı "Risk Skoru sütununa ofş'leri çarp" diyorsa, =B2*C2*D2 formülünü sadece veri olan satırlara döngüyle doğru şekilde uygula.
                - Kullanıcı "Tabloyu güzelleştir" diyorsa; başlıkları koyu yap, arkaplanı renklendir, Alignment ile ortaya daya, kolon genişliklerini (`ws.column_dimensions['A'].width = 15` gibi) otomatik ayarla.
                
                SANA SUNULAN EXCEL TABLO BAĞLAMI:
                {context}
                """
                
                try:
                    client = OpenAI(api_key=api_key)
                    response = client.chat.completions.create(
                        model="gpt-4o",
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": user_query}
                        ],
                        temperature=0
                    )
                    
                    code_string = response.choices[0].message.content.strip()
                    
                    # LLM inat edip markdown yazarsa diye güvenlik filtresi (Regex ile ayıklama):
                    import re
                    pattern = r"```(?:python)?\s*(.*?)\s*```"
                    match = re.search(pattern, code_string, re.DOTALL)
                    if match:
                        code_string = match.group(1).strip()
                    else:
                        code_string = code_string.strip()
                        
                    # İKİNCİ KATI GÜVENLİK FİLTRESİ: 
                    # Eğet AI hala sohbet metni yazdıysa ve markdown kullanmadıysa, 
                    # "import" kelimesinden önceki her şeyi kes at.
                    if "import openpyxl" in code_string:
                        code_string = code_string[code_string.find("import openpyxl"):]
                    
                    # Güvenli ortamda Çalıştırma
                    try:
                        local_env = {}
                        exec(code_string, globals(), local_env)
                        
                        success_msg = "İşlem başarıyla excelinize native(orjinal) olarak uygulandı!"
                        st.session_state["messages"].append({
                            "role": "assistant",
                            "content": success_msg,
                            "code": code_string
                        })
                        st.rerun()
                        
                    except Exception as code_error:
                        hata_metni = str(code_error)
                        st.error(f"⚠️ YAPAY ZEKA KODU YÜRÜTÜRKEN ÇÖKTÜ!\nHATA DETAYI: {hata_metni}")
                        err_msg = f"Sizin için yazdığım excel uzman kodunu çalıştırırken şu hatayı aldım:\n`{hata_metni}`\nLütfen bu hatayı ve ne yapmak istediğinizi tekrar yazar mısınız?"
                        st.session_state["messages"].append({"role": "assistant", "content": err_msg, "code": code_string})
                        
                except Exception as api_err:
                    st.error(f"OpenAI API Hatası: {str(api_err)}")
